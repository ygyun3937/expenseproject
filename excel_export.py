"""개인경비 정산 데이터를 회계 ERP 업로드용 엑셀로 내보내기.

입력: 프론트에서 전달한 JSON (작성자·개인경비 행·출장일비·유류비 등)
출력: .xlsx 바이트
"""
from __future__ import annotations

import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


CATEGORIES = ['식비', '유류비', '교통비', '주차비', '숙박비', '접대비', '소모품', '운반비', '기타']
OVERSEAS_CATEGORIES = ['식비', '교통비', '숙박비', '접대비', '소모품', '운반비', '수수료', '기타', '로밍']
ENT_RE = re.compile(r'\[접대:([^\]]+)\]\s*(.*)')


def _parse_ent_memo(memo: str) -> dict:
    """비고에서 접대비 구조 파싱: [접대:식사] 거래처 / 내부 / 외부 / 목적."""
    if not memo:
        return {}
    m = ENT_RE.search(memo)
    if not m:
        return {}
    subtype = m.group(1).strip()
    rest = m.group(2).split('|')[0].strip()
    parts = [p.strip() for p in rest.split('/') if p.strip()]
    return {
        'subtype': subtype,
        'client': parts[0] if len(parts) > 0 else '',
        'internal': parts[1] if len(parts) > 1 else '',
        'external': parts[2] if len(parts) > 2 else '',
        'purpose': parts[3] if len(parts) > 3 else '',
    }


def _parse_int(v) -> int:
    if v is None:
        return 0
    try:
        return int(str(v).replace(',', '').strip() or 0)
    except ValueError:
        return 0


def build_workbook(data: dict) -> bytes:
    wb = Workbook()
    is_overseas = data.get('tab') == 'overseas'

    # ===== Sheet 1: 개인경비 (국내/해외 분기) =====
    ws1 = wb.active
    ws1.title = '해외 개인경비' if is_overseas else '개인경비'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=('4F46E5' if is_overseas else '1E3A8A'), end_color='1E3A8A', fill_type='solid')
    thin = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if is_overseas:
        headers = ['월', '일', '통화', '외화금액', '환율', '원화금액', '카테고리', '거래처(접대)', '내부참석자', '외부참석자', '접대목적', '비고']
    else:
        headers = ['월', '일', '금액', '카테고리', '인원', '박수', '거래처(접대)', '내부참석자', '외부참석자', '접대목적', '비고']
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_num = 2
    if is_overseas:
        for row in data.get('rows', []):
            month = row.get('month', '')
            day = row.get('day', '')
            costs = row.get('costs', [])
            memo = row.get('memo', '')
            curr = row.get('currency', '')
            rate = _parse_int(row.get('rate'))
            krw_amt = _parse_int(row.get('krwAmt'))
            for idx, val in enumerate(costs):
                amt = _parse_int(val)
                if amt <= 0:
                    continue
                category = OVERSEAS_CATEGORIES[idx] if idx < len(OVERSEAS_CATEGORIES) else '기타'
                ent = _parse_ent_memo(memo) if category == '접대비' else {}
                cells = [month, day, curr, amt, rate, krw_amt, category,
                         ent.get('client', ''), ent.get('internal', ''),
                         ent.get('external', ''), ent.get('purpose', ''), memo]
                for c, v in enumerate(cells, 1):
                    cell = ws1.cell(row=row_num, column=c, value=v)
                    cell.border = border
                    if isinstance(v, (int, float)):
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right')
                row_num += 1
    else:
        for row in data.get('rows', []):
            month = row.get('month', '')
            day = row.get('day', '')
            costs = row.get('costs', [])
            memo = row.get('memo', '')
            headcount = row.get('headcount', 1)
            nights = row.get('nights', 1)
            for idx, val in enumerate(costs):
                amt = _parse_int(val)
                if amt <= 0:
                    continue
                category = CATEGORIES[idx] if idx < len(CATEGORIES) else '기타'
                ent = _parse_ent_memo(memo) if category == '접대비' else {}
                cells = [month, day, amt, category, headcount, nights,
                         ent.get('client', ''), ent.get('internal', ''),
                         ent.get('external', ''), ent.get('purpose', ''), memo]
                for c, v in enumerate(cells, 1):
                    cell = ws1.cell(row=row_num, column=c, value=v)
                    cell.border = border
                    if isinstance(v, (int, float)):
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal='right')
                row_num += 1

    # 열 너비
    widths = ([6, 6, 8, 14, 10, 14, 10, 20, 20, 22, 24, 40] if is_overseas
              else [6, 6, 14, 10, 8, 8, 20, 20, 22, 24, 40])
    for i, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ===== Sheet 2: 출장일비 =====
    ws2 = wb.create_sheet('출장일비')
    trip_headers = ['일수', '시작일', '종료일', '출장지', '금액']
    for c, h in enumerate(trip_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color='19736E', end_color='19736E', fill_type='solid')
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    for i, t in enumerate((data.get('common') or {}).get('tripRows', []) or [], 2):
        days = _parse_int(t.get('days'))
        amt = days * 30000
        for c, v in enumerate([days, t.get('startDate', ''), t.get('endDate', ''), t.get('place', ''), amt], 1):
            cell = ws2.cell(row=i, column=c, value=v)
            cell.border = border
            if isinstance(v, int):
                cell.number_format = '#,##0'

    # ===== Sheet 3: 유류비 로그 =====
    ws3 = wb.create_sheet('유류비 로그')
    fuel_headers = ['월', '일', '출발지', '도착지', '거리(km)', '통행료(원)', '내용']
    for c, h in enumerate(fuel_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color='EAB308', end_color='EAB308', fill_type='solid')
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    for i, f in enumerate((data.get('common') or {}).get('fuelRows', []) or [], 2):
        vals = [f.get('month', ''), f.get('day', ''), f.get('from', ''), f.get('to', ''),
                _parse_int(f.get('dist')), _parse_int(f.get('toll')), f.get('memo', '')]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=i, column=c, value=v)
            cell.border = border

    # ===== Sheet 4: 작성자 정보 =====
    ws4 = wb.create_sheet('작성자')
    common = data.get('common') or {}
    info = [
        ('성명', common.get('userName', '')),
        ('소속 부서', common.get('userDept', '')),
        ('작성 일자', common.get('writeDate', '')),
        ('차종', common.get('carModel', '')),
        ('차량번호', common.get('carNumber', '')),
        ('', ''),
        ('내보낸 시각', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]
    for i, (k, v) in enumerate(info, 1):
        ws4.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws4.cell(row=i, column=2, value=v)
    ws4.column_dimensions['A'].width = 15
    ws4.column_dimensions['B'].width = 30

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()
